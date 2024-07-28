import requests
import time
import sys
import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
from datetime import datetime

def obtener_fecha_hora_actual():
    """Obtiene la fecha y hora actual en formato legible."""
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def sendMessage(para, mensaje):
    url = 'http://localhost:3001/lead'  # URL de tu servidor local para la API de WhatsApp
    data = {
        "message": mensaje,
        "phone": para
    }
    headers = {
        'Content-Type': 'application/json'
    }

    max_intentos = 3
    intentos = 0
    mensaje_enviado = False

    while intentos < max_intentos:
        intentos += 1
        if intentos > 1:
            sys.stdout.write(f"\r🔄 Reintentando... ({intentos}/{max_intentos})")
            sys.stdout.flush()
        else:
            print(f"📤 [{obtener_fecha_hora_actual()}] Enviando mensaje a {para}...")

        try:
            response = requests.post(url, json=data, headers=headers, timeout=10)
            response.raise_for_status()  # Lanza una excepción para errores HTTP

            if response.status_code == 404:
                sys.stdout.write(f"\r🔍 [{obtener_fecha_hora_actual()}] El número {para} no tiene WhatsApp. Pasando al siguiente número.\n")
                sys.stdout.flush()
                return False  # Indicar que no se pudo enviar el mensaje
            else:
                print(f"✅ [{obtener_fecha_hora_actual()}] Mensaje enviado con éxito a {para}.")
                mensaje_enviado = True
                break  # Salir del bucle si el mensaje fue enviado con éxito

        except requests.exceptions.Timeout:
            if intentos == max_intentos:
                # Mensaje final después de todos los intentos fallidos
                sys.stdout.write(f"\r❌ [{obtener_fecha_hora_actual()}] El mensaje a {para} no pudo ser enviado después de {max_intentos} intentos. El número de teléfono no posee WhatsApp.\n")
                sys.stdout.flush()
            else:
                sys.stdout.write(f"\r🔄 Reintentando... ({intentos}/{max_intentos})")
                sys.stdout.flush()
            time.sleep(5)  # Esperar antes de reintentar
        except requests.exceptions.RequestException as req_err:
            print(f"🚨 [{obtener_fecha_hora_actual()}] Error de solicitud al enviar mensaje a {para}: {req_err}")
            break  # No continuar intentando en caso de otros errores de solicitud
        except Exception as e:
            print(f"⚠️ [{obtener_fecha_hora_actual()}] Error inesperado al enviar mensaje a {para}: {e}")
            break  # No continuar intentando en caso de errores inesperados

    return mensaje_enviado  # Indicar si el mensaje fue enviado con éxito o no

def select_excel_file():
    """Abre un cuadro de diálogo para seleccionar un archivo Excel."""
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal de Tkinter
    file_path = filedialog.askopenfilename(
        title="Selecciona el archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    return file_path

def main():
    excel_file = select_excel_file()
    
    if not excel_file:
        print("❌ No se seleccionó ningún archivo Excel.")
        return

    if not os.path.exists(excel_file):
        print(f"❌ El archivo '{excel_file}' no se encontró.")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        has_invalid_data = False

         # Iterar sobre las filas del archivo Excel
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Empezar desde la segunda fila
            # Asegurarse de que los datos no estén vacíos antes de procesarlos
            numero = str(row[0]).strip().replace('-', '') if row[0] else ''
            nombre_estudiante = str(row[2]).strip() if row[2] else ''
            nombre_tecnico = str(row[3]).strip() if row[3] else ''
            falla_reportada = str(row[4]).strip() if row[4] else ''
            serie = str(row[6]).strip() if row[6] else ''


             # Verificar que no haya valores vacíos
            if not numero or not nombre_estudiante or not nombre_tecnico or not falla_reportada or not serie:
                print(f"❌ Alguno de los valores en la fila {sheet.iter_rows.index(row) + 1} no es válido o está vacío.")
                continue  # Omitir esta fila

            # Agregar el prefijo del país si no está presente
            if len(numero) == 8:  # Asumiendo que los números sin prefijo tienen 8 dígitos
                numero = '503' + numero

            mensaje = (
                f"🌟 *¡Hola!* Nos comunicamos con el responsable del estudiante:\n"
                f"\n📋 *Datos del ESTUDIANTE*: {nombre_estudiante}\n"
                f"\n📱 Desde el Ministerio de Educación, necesitamos que vengas con el equipo que has reportado en nuestra lista de espera con la siguiente falla: *{falla_reportada}* - Serie del equipo: *{serie}.*\n"
                f"\n🔧💻 Procederemos a solucionar tu caso y reparar el equipo.\n"
                f"\n📲 Por favor, presenta este mensaje y tu DUI al llegar para agilizar el proceso.\n"
                f"\n⏰ Nuestros horarios de atención son de Lunes a Viernes, de 7:30 am a 12:00 md y de 1:00 pm a 3:30 pm.\n"
                f"\n📍 *Dirección*: Estamos ubicados a solo una cuadra del ISSS Santa Ana, en la Avenida California, colonia El Palmar, dentro de CE INSA Industrial. [Ver ubicación](https://maps.app.goo.gl/9G4GNo8RE63VKH6r6)\n"
                f"\n🔌 Por favor, asegúrate de traer el equipo completo.\n"
                f"\n🧑‍🔧 Técnico a cargo: *{nombre_tecnico}*"
            )

            sendMessage(numero, mensaje)

            time.sleep(10)

    except Exception as e:
        print(f"⚠️ Error inesperado al procesar el archivo Excel: {e}")

    input("Presiona Enter para cerrar...")

if __name__ == "__main__":
    main()
